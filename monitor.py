import os
import getpass
import time
from datetime import datetime
import threading
import pygetwindow as gw
from openpyxl import Workbook, load_workbook
from pynput import mouse, keyboard
from threading import Lock

last_activity_time = datetime.now()
last_window_title = None
idle_start_time = None
current_activity_start_time = datetime.now()
current_date = datetime.now().date()
last_backup_time = datetime.now()

idle_threshold = 60  # Start idling after - (seconds)

backup_interval_in_hours = 0.50  # Backup interval in hours (0.50 = 30mins)
backup_location = os.path.join(os.path.expanduser("~"), "Documents", "Activity_Backup")

documents_folder = os.path.join(os.path.expanduser("~"), "Documents")
activity_folder = os.path.join(documents_folder, "Activity")

excel_lock = Lock()

# Sleep interval for the monitoring loop
sleep_interval = 1  # Sleep interval in seconds


def get_log_path():
    file_name = f"{getpass.getuser()}_{current_date.strftime('%Y-%m-%d')}.xlsx"
    return os.path.join(activity_folder, file_name)


def get_backup_log_path():
    log_path = get_log_path()
    base_name, ext = os.path.splitext(os.path.basename(log_path))
    backup_file_name = f"{base_name}_backup{ext}"
    return os.path.join(backup_location, backup_file_name)


def ensure_activity_folder():
    if not os.path.exists(activity_folder):
        os.makedirs(activity_folder)


def ensure_backup_folder():
    if not os.path.exists(backup_location):
        os.makedirs(backup_location)


def initialize_workbook():
    ensure_activity_folder()

    log_path = get_log_path()

    if not os.path.exists(log_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Activity Log"

        headers = ["LoginName", "MOTION_APPLICATION_CR", "MOTION_TYPE_CR", "START_TIME_DT", "END_TIME_DT", "TotalTime"]
        for col_num, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_num).value = header

        wb.create_sheet(title="Summary")
        wb["Summary"].cell(row=1, column=1).value = "Idle Hours"
        wb.save(log_path)
    return load_workbook(log_path)


def update_activity():
    global last_activity_time
    last_activity_time = datetime.now()


def get_active_window():
    try:
        active_window = gw.getActiveWindow()
        if active_window:
            window_title = active_window.title
            return f"{window_title}"
    except Exception as e:
        print(f"Error getting active window: {e}")
        return "Unknown Window"
    return "Unknown Window"


def log_to_excel(activity_type, window_title, start_time, end_time):
    with excel_lock:
        wb = initialize_workbook()
        ws = wb.active

        next_row = ws.max_row + 1

        # Populate columns based on format
        ws.cell(row=next_row, column=1).value = getpass.getuser()
        ws.cell(row=next_row, column=2).value = window_title
        ws.cell(row=next_row, column=3).value = activity_type
        ws.cell(row=next_row, column=4).value = start_time.strftime('%Y-%m-%d %H:%M:%S')
        ws.cell(row=next_row, column=5).value = end_time.strftime('%Y-%m-%d %H:%M:%S')

        total_time = end_time - start_time
        total_seconds = int(total_time.total_seconds())
        hours, remainder = divmod(total_seconds, 3600)
        minutes, seconds = divmod(remainder, 60)
        formatted_time = f"{hours}:{minutes:02}:{seconds:02}"  # h:mm:ss format
        ws.cell(row=next_row, column=6).value = formatted_time

        # Try to save the workbook, handling the case where the file might be locked
        max_retries = 5
        retries = 0
        while retries < max_retries:
            try:
                wb.save(get_log_path())
                break
            except PermissionError:
                retries += 1
                print(f"Unable to save {get_log_path()}. It might be open. Retrying in 5 seconds...")
                time.sleep(5)
        else:
            print(f"Failed to save {get_log_path()} after {max_retries} attempts.")


def on_click(x, y, button, pressed):
    if pressed:
        update_activity()


def on_move(x, y):
    update_activity()


def on_press(key):
    update_activity()


def backup_excel_file():
    ensure_backup_folder()
    try:
        log_path = get_log_path()
        backup_file_path = get_backup_log_path()
        import shutil
        shutil.copyfile(log_path, backup_file_path)
        print(f"Backup created at {backup_file_path}")
    except Exception as e:
        print(f"Failed to create backup: {e}")


def monitor_activity():
    global last_activity_time, idle_start_time, current_activity_start_time, last_window_title, current_date, last_backup_time

    while True:
        current_time = datetime.now()

        time_since_last_activity = (current_time - last_activity_time).total_seconds()

        if current_time.date() != current_date:
            # Handle the end of the day
            end_of_day = datetime.combine(current_date, datetime.max.time())
            if idle_start_time:
                log_to_excel("Idle", "Idle Hours", idle_start_time, end_of_day)
                idle_start_time = None
            elif last_window_title:
                log_to_excel("Working", last_window_title, current_activity_start_time, end_of_day)
                last_window_title = None

            current_date = current_time.date()
            initialize_workbook()  # Create a new workbook for the new day
            current_activity_start_time = datetime.combine(current_date, datetime.min.time())

        if time_since_last_activity < idle_threshold:
            if idle_start_time:
                log_to_excel("Idle", "Idle Hours", idle_start_time, current_time)
                idle_start_time = None

            active_window = get_active_window()
            if active_window != last_window_title:
                if last_window_title:
                    log_to_excel("Working", last_window_title, current_activity_start_time, current_time)

                current_activity_start_time = current_time
                last_window_title = active_window
        else:
            if not idle_start_time:
                idle_start_time = current_time
                if last_window_title:
                    log_to_excel("Working", last_window_title, current_activity_start_time, current_time)
                    last_window_title = None

        # Check if it's time to back up the Excel file
        if (current_time - last_backup_time).total_seconds() >= backup_interval_in_hours * 3600:
            backup_excel_file()
            last_backup_time = current_time

        time.sleep(sleep_interval)  # Check every 1 second


if __name__ == "__main__":
    try:
        initialize_workbook()

        mouse_listener = mouse.Listener(on_click=on_click, on_move=on_move)
        keyboard_listener = keyboard.Listener(on_press=on_press)

        mouse_listener.start()
        keyboard_listener.start()

        activity_thread = threading.Thread(target=monitor_activity)
        activity_thread.daemon = True
        activity_thread.start()

        mouse_listener.join()
        keyboard_listener.join()
    except KeyboardInterrupt:
        print("Shutting down gracefully.")
        mouse_listener.stop()
        keyboard_listener.stop()
